import streamlit as st
from PIL import Image
from datetime import date
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from sqlalchemy import create_engine, Table, MetaData, Column, Integer, String, Boolean, Date

# --- Configuración de Streamlit ---
st.set_page_config(page_title="Checklist Área de Planificación", page_icon="✅")
img = Image.open("logo.png")

# --- Conexión a Supabase ---
DATABASE_URL = "postgresql://postgres:Planificaci%C3%B3nretail2025@db.xxxx.supabase.co:5432/postgres"
engine = create_engine(DATABASE_URL)

# --- Definir tabla manualmente (sin autoload) ---
metadata = MetaData()
checklist_table = Table(
    'checklist', metadata,
    Column('id', Integer, primary_key=True),
    Column('tienda', String(100)),
    Column('encargado', String(100)),
    Column('fecha', Date),
    Column('tarea', String(100)),
    Column('completada', Boolean),
    Column('valor', String(20)),
    Column('comentario', String(500))
)

# --- Título con imagen al lado ---
col1, col2 = st.columns([0.1, 1])
col1.image(img, width=60)
col2.markdown("## Checklist Área de Planificación")

# --- Selección de tienda, encargado y fecha ---
col1, col2, col3 = st.columns(3)
tiendas = [
    "Plaza Oeste","Florida Center","Plaza Alameda","Plaza Sur","Portal Rancagua","Plaza Trebol",
    "Plaza Vespucio","Plaza Los Angeles","Plaza Norte","Apumanque","Portal Ñuñoa","Plaza Calama",
    "Plaza Antofagasta","Portal Temuco","Arauco Premium Outlet","Paseo Viña Centro","Costanera Center",
    "Plaza Bio Bio","Arauco Maipu","Mall del Centro Concepcion","Plaza Tobalaba","Portal Osorno",
    "Arauco Quilicura","Plaza Copiapo","Plaza Egaña","Open Plaza Ovalle","Easton Quilicura",
    "Vivo Outlet Maipú","Plaza Iquique","Vivo Los Trapenses","Gran Avenida Esp.Urbano",
    "San Pedro Arauco Outlet","Portal Centro Talca","Mall Valle Curico","Arauco Chillan","Midmall Outlet",
    "Paseo Estado","Paseo Chiloé","La Fábrica Patio Outlet","Vivo Outlet Peñuelas","Viña Outlet Park",
    "Valdivia Centro","Vivo San Fernando","Parque Arauco","Vivo Outlet La Florida","Portal Valdivia",
    "Patio Rancagua","Vivo Coquimbo","Vivo Outlet Temuco","Arauco Coronel","Mall Barrio Independencia",
    "Outlet Style","Easton Temuco","PATRONATO 403","Portal El Llano","Paseo Alerce","Santa Filomena 540",
    "Antofagasta Outlet Espacio Urbano","Paseo Costanera Puerto Montt","Paseo Puerto Varas",
    "Vivo Outlet Chillan","Curauma Outlet","15 Norte Viña","Talca Outlet Go Florida","Pionero Punta Arenas",
    "Rancagua Outlet Mall","Mall Paseo San Bernardo","Paseo Quillota","Easton Segundo Piso"
]
tienda = col1.selectbox("🏪 Tienda", tiendas)

encargados = ["Brany Gómez", "Gerardo Muñoz", "Juan Pablo"]
encargado = col2.selectbox("👤 Encargado", encargados)

fecha_checklist = col3.date_input("📅 Fecha del checklist", value=date.today())

st.markdown("---")

# --- Lista de tareas ---
st.subheader("Puntos a revisar")
tareas = [
    "Cubicación vestuario",
    "Cubicación calzado",
    "Reposición (Curva, RAMI)",
    "Despachos",
    "Club Pillin",
    "Mix colección",
    "Visual merchandising",
    "Competencia",
    "Gestión equipo de venta",
    "Dotación",
    "Experiencia del cliente (CX)",
    "Posibles áreas de mejora"
]

estado = []
valores_comentario = []
valores_opcion = []
opcion_cub = ["70%","75%","80%","85%","90%","95%","100%","105%","110%","115%","120%","125%","130%"]
dot = [1,2,3,4,5,6]

for tarea in tareas:
    checked = st.checkbox(tarea, key=f"chk_{tarea}")
    estado.append(checked)
    
    if tarea in ["Cubicación vestuario", "Cubicación calzado"]:
        if tarea == "Cubicación vestuario":
            opciones = st.selectbox(f"Porcentaje vestuario", opcion_cub, index=6, key=f"opt_{tarea}")
        else:
            opciones = st.selectbox(f"Porcentaje calzado", opcion_cub, index=6, key=f"opt_{tarea}")
        valores_opcion.append(opciones)
        comentario = st.text_input(f"Comentario para {tarea}", key=f"com_{tarea}") if checked else ""
        valores_comentario.append(comentario)
        
    elif tarea == "Dotación":
        opciones = st.selectbox(f"Cantidad de personal", dot, index=2, key=f"opt_{tarea}")
        valores_opcion.append(opciones)
        comentario = st.text_input(f"Comentario para {tarea}", key=f"com_{tarea}") if checked else ""
        valores_comentario.append(comentario)
        
    else:
        valores_opcion.append("")
        comentario = st.text_input(f"Comentario para {tarea}", key=f"com_{tarea}") if checked else ""
        valores_comentario.append(comentario)

# --- Progreso ---
completadas = sum(estado)
total = len(tareas)
progreso = completadas / total if total > 0 else 0

st.progress(progreso)
st.write(f"Haz completado **{completadas} de {total} ítems**.")

faltantes = total - completadas
if completadas == total:
    st.success("🎉 ¡Checklist completo!")
elif completadas > 0:
    st.info(f"💪 Aún faltan **{faltantes}** puntos por abordar." )
else:
    st.warning("🙌 Aún no comienzas tu Checklist")

# --- Botón para guardar en DB y descargar Excel ---
if all(estado):
    if st.button("✅ Completado"):
        fecha_str = fecha_checklist.strftime("%d-%m-%Y")

        # Crear DataFrame
        df = pd.DataFrame({
            "Tarea": tareas,
            "Completada": estado,
            "Valor": valores_opcion,
            "Comentario": valores_comentario
        })    

        # Guardar a Excel en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Checklist", startrow=3)

        # Formato con openpyxl
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active

        ws["A1"] = f"Tienda: {tienda}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws["A2"] = f"Encargado: {encargado}"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        ws["A3"] = f"Fecha: {fecha_str}"
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border

        for row in ws.iter_rows(min_row=4, max_row=3+len(df)+1, min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        for cell in ws[4]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Guardar cambios Excel
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        # Botón descarga
        st.download_button(
            label="📥 Descargar checklist",
            data=final_output,
            file_name="Checklist_Completo.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # --- Guardar datos en Supabase ---
        with engine.connect() as conn:
            for i, tarea in enumerate(tareas):
                conn.execute(checklist_table.insert().values(
                    tienda=tienda,
                    encargado=encargado,
                    fecha=fecha_checklist,
                    tarea=tarea,
                    completada=estado[i],
                    valor=valores_opcion[i],
                    comentario=valores_comentario[i]
                ))
else:
    st.error("❌ Debes marcar todos los check antes de completar el checklist.")
















